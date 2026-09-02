"""Smart Import for AlphaX CRM.

List-view driven CSV/XLSX import. Pre-scans every Link field on the chosen
target doctype for master records that don't exist yet (Lead Source,
Territory, Industry Type, AlphaX PreLead Status, ...) and lets the user
decide per master doctype: auto-create the missing records or skip (blank
the field) — then imports row-by-row under savepoints so one bad row never
kills the batch.

Ported from the 0.4.x "Smart Import" feature (originally Prospect/Lead only,
Prospect since renamed to AlphaX PreLead)
to the current schema, which added Link-typed fields to AlphaX PreLead
(status/source/territory/industry) and introduced AlphaX Smart Lead as a
third, canonical data-entry target that auto-maps into Lead on insert. All
three targets are supported here:

    AlphaX PreLead   — calling list; operator qualifies to Lead later
    Lead              — created directly
    AlphaX Smart Lead — canonical entry; syncs itself into a Lead on insert

Wired from the AlphaX PreLead list view (alphax_prelead_list.js). Files
above SYNC_ROW_LIMIT rows are enqueued on the long queue and the user is
notified over realtime when the job completes. All inserts run under the
caller's own permissions — no ignore_permissions anywhere in this module.
"""

import csv
import io

import frappe
from frappe import _
from frappe.utils import cint

SKIP_FIELDTYPES = {
    "Table", "Table MultiSelect", "Section Break", "Column Break",
    "Tab Break", "HTML", "Button", "Fold", "Heading",
}
SYNC_ROW_LIMIT = 500

# Doctypes this feature is allowed to create, and the permission label used
# when talking to the user. Keep in sync with the list view's target picker.
TARGET_DOCTYPES = ("AlphaX PreLead", "Lead", "AlphaX Smart Lead")
DEFAULT_TARGET = "AlphaX PreLead"

# Name fields to try, in priority order, when deciding how a target doctype
# wants a person's name: "split" means first_name + last_name; anything else
# is a single fieldname to hold the full name.
NAME_FIELD_PRIORITY = ("first_name", "prospect_name", "lead_name", "contact_name")


# ---------------------------------------------------------------------------
# File reading + header mapping
# ---------------------------------------------------------------------------
def _read_rows(file_url: str) -> list[dict]:
    fdoc = frappe.get_doc("File", {"file_url": file_url})
    content = fdoc.get_content()
    if isinstance(content, str):
        content = content.encode("utf-8")
    fname = (fdoc.file_name or "").lower()
    if fname.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        grid = [
            ["" if c is None else str(c).strip() for c in row]
            for row in ws.iter_rows(values_only=True)
        ]
    else:
        text = content.decode("utf-8-sig", errors="replace")
        sample = text[:4096]
        delim = "\t" if sample.count("\t") > sample.count(",") else ","
        grid = list(csv.reader(io.StringIO(text), delimiter=delim))
    grid = [r for r in grid if any(str(c).strip() for c in r)]
    if len(grid) < 2:
        frappe.throw(_("File has no data rows"))
    headers = [str(h).strip() for h in grid[0]]
    return [{h: str(c).strip() for h, c in zip(headers, r)} for r in grid[1:]]


def _field_map(headers: list[str], doctype: str):
    """Map file headers to docfields by fieldname, label, or normalized form.

    Tries an exact fieldname match, then an exact label match, then a
    normalized match (case/spacing/punctuation-insensitive — "Email Id",
    "email-id", "EMAIL_ID" and "email id" all resolve to the field
    "email_id") before giving up. This catches formatting differences
    automatically; genuinely different wording (e.g. "Work Email" for
    email_id) still needs an entry in FIELD_ALIASES below, since that's a
    vocabulary difference, not a formatting one.
    """
    meta = frappe.get_meta(doctype)
    by_fieldname = {df.fieldname: df for df in meta.fields}
    by_label = {(df.label or "").strip().lower(): df for df in meta.fields if df.label}
    by_normalized = {}
    for df in meta.fields:
        for key in (df.fieldname, df.label or ""):
            norm = _normalize_header(key)
            if norm:
                by_normalized.setdefault(norm, df)
    mapped, unmapped = {}, []
    for h in headers:
        df = by_fieldname.get(h) or by_label.get(h.lower()) or by_normalized.get(_normalize_header(h))
        if df and df.fieldtype not in SKIP_FIELDTYPES:
            mapped[h] = df
        else:
            unmapped.append(h)
    return mapped, unmapped


def _normalize_header(s: str) -> str:
    import re

    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def _name_mode(meta) -> str | None:
    """How does this target doctype want a person's name?

    Returns "split" (first_name/last_name), a single fieldname to hold the
    full name, or None if the doctype has none of the recognized fields.
    """
    fieldnames = {df.fieldname for df in meta.fields}
    if "first_name" in fieldnames:
        return "split"
    for f in NAME_FIELD_PRIORITY[1:]:
        if f in fieldnames:
            return f
    return None


# ---------------------------------------------------------------------------
# Source-format normalization
# ---------------------------------------------------------------------------
# Recognized external export headers, per canonical target field, in fallback
# priority order (first non-empty value wins). Covers the contact-export
# format customers receive by email ("Work email", "Contact name", ...), and
# both spellings a target doctype might use for the same concept (e.g.
# company name is "company_name" on PreLead/Lead but "organization" on
# Smart Lead — harmless no-op for whichever target doesn't have that field).
# Extend this table to onboard new export formats — no other change needed.
FIELD_ALIASES = {
    "email_id": ("work email", "work email 2", "private email", "email address"),
    "email": ("work email", "work email 2", "private email", "email address"),
    "mobile_no": ("mobile", "mobile 2", "mobile number"),
    "phone": ("phone", "phone 2", "direct phone", "direct phone 2"),
    "company_name": ("company name", "company", "organization"),
    "organization": ("company name", "company", "organization"),
    "job_title": ("job title", "title", "designation"),
    "website": ("linkedin profile", "linkedin", "website url"),
    "city": ("contact location - city", "location city"),
    "industry": ("industry",),
}
NAME_ALIASES = (
    "contact name", "full name", "name",
    # Raw fieldname-style headers from exports generated by another AlphaX
    # doctype (e.g. a Lead-shaped export dropped onto a PreLead import) —
    # these are the source doctypes' own name fields, so they only get
    # picked up here as a fallback when they don't already direct-map onto
    # the current target.
    "lead_name", "prospect_name", "prospect name", "contact_name",
)
PHONE_FIELDS = ("mobile_no", "phone", "whatsapp_no")


def _normalize_phone(value: str) -> str:
    import re

    v = re.sub(r"[^\d+]", "", value or "")
    if not v:
        return ""
    if v.startswith("00"):
        v = "+" + v[2:]
    if v.startswith("966"):
        v = "+" + v
    elif v.startswith("05") and len(v) == 10:
        v = "+966" + v[1:]
    return v


def _transform_rows(rows: list[dict], target: str = DEFAULT_TARGET):
    """Normalize external export rows into target-fieldname rows.

    Merges alias columns (e.g. Work email -> Work email 2 -> Private email)
    into their canonical target field, first non-empty value winning; a
    column that already direct-maps to the field takes top priority. Fills
    the target's name field(s) from a "Contact name" style column and
    normalizes phone numbers. Idempotent: an already-import-ready file passes
    through unchanged. Returns (rows, derived) where derived describes what
    was auto-mapped, for display in the confirmation dialog.
    """
    if not rows:
        return rows, {}
    meta = frappe.get_meta(target)
    headers = list(rows[0].keys())
    mapped, _unmapped = _field_map(headers, target)
    # Normalized lookup (case/spacing/punctuation-insensitive) so alias
    # entries match "Work_Email", "WorkEmail", "work-email", etc., not just
    # the exact "work email" spelling.
    lower = {_normalize_header(h): h for h in headers}
    fieldname_to_header = {}
    for h, df in mapped.items():
        fieldname_to_header.setdefault(df.fieldname, h)

    derived, plan, consumed = {}, {}, set()
    for fieldname, aliases in FIELD_ALIASES.items():
        if not meta.has_field(fieldname):
            continue
        sources = []
        direct_h = fieldname_to_header.get(fieldname)
        if direct_h:
            sources.append(direct_h)
        alias_hs = [
            lower[_normalize_header(a)]
            for a in aliases
            if _normalize_header(a) in lower and lower[_normalize_header(a)] not in sources
        ]
        if not alias_hs:
            continue
        sources.extend(alias_hs)
        plan[fieldname] = sources
        consumed.update(sources)
        derived[fieldname] = sources

    name_mode = _name_mode(meta)
    name_header = None
    already_named = (
        any(f in fieldname_to_header for f in ("first_name", "last_name"))
        if name_mode == "split"
        else (name_mode and name_mode in fieldname_to_header)
    )
    if name_mode and not already_named:
        for a in NAME_ALIASES:
            if _normalize_header(a) in lower:
                name_header = lower[_normalize_header(a)]
                consumed.add(name_header)
                label = "first_name / last_name" if name_mode == "split" else name_mode
                derived[label] = [name_header]
                break

    if not plan and not name_header:
        for r in rows:
            for f in PHONE_FIELDS:
                h = fieldname_to_header.get(f)
                if h and r.get(h):
                    r[h] = _normalize_phone(r[h])
        return rows, {}

    passthrough = [h for h in mapped if h not in consumed]
    out = []
    for r in rows:
        new = {h: r.get(h, "") for h in passthrough}
        for fieldname, sources in plan.items():
            new[fieldname] = next((r[s].strip() for s in sources if (r.get(s) or "").strip()), "")
        if name_header:
            full = (r.get(name_header) or "").strip()
            if name_mode == "split":
                parts = full.split()
                new["first_name"] = parts[0] if parts else ""
                new["last_name"] = " ".join(parts[1:])
            else:
                new[name_mode] = full
        for f in PHONE_FIELDS:
            key = f if f in new else fieldname_to_header.get(f)
            if key and new.get(key):
                new[key] = _normalize_phone(new[key])
        out.append(new)
    return out, derived


# ---------------------------------------------------------------------------
# Permissions
# ---------------------------------------------------------------------------
def _target_permissions() -> dict:
    return {dt: bool(frappe.has_permission(dt, "create")) for dt in TARGET_DOCTYPES}


def _check_target(target: str, perms: dict):
    if target not in TARGET_DOCTYPES:
        frappe.throw(_("Unsupported import target: {0}").format(target))
    if not perms.get(target):
        frappe.throw(_("Not permitted to create {0}").format(_(target)), frappe.PermissionError)


# ---------------------------------------------------------------------------
# Pre-scan  (what would this file need, for the chosen target?)
# ---------------------------------------------------------------------------
@frappe.whitelist()
def scan_file(file_url: str, import_as: str = DEFAULT_TARGET):
    perms = _target_permissions()
    if not any(perms.values()):
        frappe.throw(_("Not permitted"), frappe.PermissionError)
    if import_as not in perms or not perms.get(import_as):
        import_as = next((dt for dt in TARGET_DOCTYPES if perms.get(dt)), None)
    _check_target(import_as, perms)

    rows = _read_rows(file_url)
    rows, derived = _transform_rows(rows, import_as)
    mapped, unmapped = _field_map(list(rows[0].keys()), import_as)
    missing = {}
    for header, df in mapped.items():
        if df.fieldtype != "Link":
            continue
        values = {r.get(header, "") for r in rows} - {""}
        absent = sorted(v for v in values if not frappe.db.exists(df.options, v))
        if absent:
            entry = missing.setdefault(df.options, {"values": set(), "fields": []})
            entry["values"].update(absent)
            entry["fields"].append(df.label or df.fieldname)
    return {
        "total_rows": len(rows),
        "import_as": import_as,
        "available_targets": [dt for dt in TARGET_DOCTYPES if perms.get(dt)],
        "mapped_fields": [df.label or df.fieldname for df in mapped.values()],
        "unmapped_columns": unmapped,
        "derived_fields": {k: v for k, v in derived.items()},
        "missing_masters": {
            dt: {"values": sorted(v["values"]), "fields": v["fields"]}
            for dt, v in missing.items()
        },
    }


# ---------------------------------------------------------------------------
# Master auto-creation  (generic across simple + tree masters)
# ---------------------------------------------------------------------------
def _create_master(doctype: str, value: str):
    meta = frappe.get_meta(doctype)
    doc = frappe.new_doc(doctype)
    autoname = (meta.autoname or "").strip()
    if autoname.startswith("field:"):
        doc.set(autoname.split(":", 1)[1], value)
    elif meta.title_field:
        doc.set(meta.title_field, value)
        doc.name = value
    else:
        doc.name = value
    if meta.is_tree:
        parent_field = f"parent_{frappe.scrub(doctype)}"
        root = frappe.db.get_value(doctype, {parent_field: ("in", ("", None))}, "name")
        if root:
            doc.set(parent_field, root)
    for df in meta.fields:
        if df.reqd and not doc.get(df.fieldname) and df.fieldtype in ("Data", "Small Text", "Text"):
            doc.set(df.fieldname, value)
    doc.insert()
    return doc.name


def _is_duplicate(data: dict) -> bool:
    email_value = data.get("email_id") or data.get("email")
    mobile_value = data.get("mobile_no")
    if not email_value and not mobile_value:
        return False
    # Each target doctype spells "email" differently — build filters per
    # doctype rather than reusing one filter list, so the WHERE clause never
    # references a column a given doctype doesn't actually have.
    email_field_by_doctype = {
        "Lead": "email_id",
        "AlphaX PreLead": "email_id",
        "AlphaX Smart Lead": "email",
    }
    for dt, email_field in email_field_by_doctype.items():
        or_filters = []
        if email_value:
            or_filters.append([email_field, "=", email_value])
        if mobile_value:
            or_filters.append(["mobile_no", "=", mobile_value])
        if or_filters and frappe.get_all(dt, or_filters=or_filters, limit=1):
            return True
    return False


# ---------------------------------------------------------------------------
# Import  (sync for small files, long queue beyond SYNC_ROW_LIMIT)
# ---------------------------------------------------------------------------
@frappe.whitelist()
def run_import(file_url: str, decisions=None, skip_duplicates=1, run_ai=0, default_source=None,
               import_as=DEFAULT_TARGET):
    perms = _target_permissions()
    _check_target(import_as, perms)
    decisions = frappe.parse_json(decisions or "{}")
    rows = _read_rows(file_url)
    if len(rows) > SYNC_ROW_LIMIT:
        frappe.enqueue(
            "alphax_crm.api.lead_import._execute",
            queue="long",
            file_url=file_url,
            decisions=decisions,
            skip_duplicates=cint(skip_duplicates),
            run_ai=cint(run_ai),
            user=frappe.session.user,
            notify=True,
            default_source=default_source,
            target=import_as,
        )
        return {"queued": True, "total_rows": len(rows)}
    return _execute(
        file_url, decisions, cint(skip_duplicates), cint(run_ai),
        frappe.session.user, default_source=default_source, target=import_as,
    )


def _execute(file_url, decisions, skip_duplicates, run_ai, user, notify=False,
             default_source=None, target=DEFAULT_TARGET):
    rows = _read_rows(file_url)
    rows, _derived = _transform_rows(rows, target)
    mapped, _unmapped = _field_map(list(rows[0].keys()), target)

    created_masters = {}
    for header, df in mapped.items():
        if df.fieldtype != "Link":
            continue
        action = decisions.get(df.options, "skip")
        values = {r.get(header, "") for r in rows} - {""}
        for v in sorted(values):
            if frappe.db.exists(df.options, v):
                continue
            if action == "create":
                _create_master(df.options, v)
                created_masters.setdefault(df.options, []).append(v)
            else:
                for r in rows:
                    if r.get(header) == v:
                        r[header] = ""

    batch = frappe.utils.now_datetime().strftime("IMP-%Y%m%d-%H%M%S")
    prospect_meta = frappe.get_meta("AlphaX PreLead")
    inserted, skipped, failed = [], [], []
    for idx, r in enumerate(rows, start=2):
        data = {df.fieldname: r[h] for h, df in mapped.items() if r.get(h)}
        if not data:
            continue
        if default_source:
            source_field = "lead_source" if target == "AlphaX Smart Lead" else "source"
            if not data.get(source_field):
                data[source_field] = default_source
        if skip_duplicates and _is_duplicate(data):
            skipped.append({"row": idx, "reason": _("Duplicate (email/mobile match)")})
            continue
        frappe.db.savepoint("alphax_lead_import_row")
        try:
            doc = frappe.new_doc(target)
            doc.update(data)
            if target == "Lead":
                doc.flags.alphax_skip_ai = not run_ai
            elif target == "AlphaX PreLead" and prospect_meta.has_field("import_batch"):
                doc.import_batch = batch
            doc.insert()
            inserted.append(doc.name)
        except Exception as e:
            frappe.db.rollback(save_point="alphax_lead_import_row")
            failed.append({"row": idx, "reason": str(e)})
    frappe.db.commit()

    summary = {
        "inserted": len(inserted),
        "skipped": skipped,
        "failed": failed,
        "created_masters": created_masters,
        "target": target,
    }
    if notify:
        frappe.publish_realtime("alphax_lead_import_done", summary, user=user, after_commit=True)
    return summary
